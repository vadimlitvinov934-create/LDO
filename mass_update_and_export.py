import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from werkzeug.security import generate_password_hash
from models import SessionLocal, User, Student

# ==========================================
# 1. СПИСКИ ПОЛЬЗОВАТЕЛЕЙ И НОВЫХ ПАРОЛЕЙ
# ==========================================

# Сотрудники
STAFF = [
    {"role": "head",     "fio": "Иванова Галина Петровна",        "pass": "HEADPO"},
    {"role": "curator",  "fio": "Султангазинова Диана Сериковна", "pass": "SDS2025PO115"},
    {"role": "curator",  "fio": "Брусенко Владислав Сергеевич",   "pass": "BRU2025PO175"},
    {"role": "starosta", "fio": "Староста ПО175",                 "pass": "STAR175"},
    
    # НОВЫЙ АККАУНТ:
    {"role": "tech",     "fio": "Техническая Поддержка",          "pass": "tech123"},
]

# Студенты (ФИО: Пароль)
STUDENTS_PASS = {
    "Агулярный Мирослав Николаевич": "mirA24x",
    "Ашимбеков Алибек Ерланбекович": "ashEr09",
    "Бимурзин Алишер Асхатович": "bimAli27",
    "Литвинов Вадим Игоревич": "lv2009",
    # ... добавь сюда остальных студентов ...
}

# ==========================================
# 2. СКРИПТ ОБНОВЛЕНИЯ И ЭКСПОРТА
# ==========================================

def main():
    # Создаем папку passwords, если её нет
    output_dir = "passwords"
    os.makedirs(output_dir, exist_ok=True)

    # Имя файла с датой и временем
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
    filename = f"{output_dir}/new_passwords_{timestamp}.xlsx"

    # Настраиваем Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Доступы"
    
    # Заголовки
    headers = ["ФИО", "Роль / Группа", "Логин (ФИО)", "Пароль"]
    ws.append(headers)

    # Стили для красоты
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Красим шапку
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    session = SessionLocal()
    count = 0

    print("--- Обновление сотрудников ---")
    for u in STAFF:
        # Обновляем БД
        user_db = session.query(User).filter_by(fio=u["fio"]).first()
        p_hash = generate_password_hash(u["pass"])
        
        if not user_db:
            user_db = User(username=u["fio"], fio=u["fio"], role=u["role"], password_hash=p_hash)
            session.add(user_db)
        else:
            user_db.password_hash = p_hash
            user_db.role = u["role"]
        
        # Пишем в Excel
        ws.append([u["fio"], u["role"], u["fio"], u["pass"]])
        count += 1
        print(f"✔️ {u['fio']}")

    print("\n--- Обновление студентов ---")
    for fio, raw_pass in STUDENTS_PASS.items():
        st_db = session.query(Student).filter_by(full_name=fio).first()
        
        group_info = "Студент"
        if st_db:
            # Обновляем пароль в БД
            st_db.password_hash = generate_password_hash(raw_pass)
            if st_db.group_code:
                group_info = st_db.group_code
            
            # Пишем в Excel
            ws.append([fio, group_info, fio, raw_pass])
            count += 1
            print(f"✔️ {fio}")
        else:
            # Если студента нет в базе, просто пишем в Excel пометку (или пропускаем)
            ws.append([fio, "НЕТ В БАЗЕ", fio, raw_pass])
            print(f"⚠️ {fio} — нет в базе данных!")

    # Сохраняем изменения в БД
    session.commit()
    session.close()

    # Автоширина колонок в Excel
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
        
        # Рамки всем ячейкам
        for cell in col:
            cell.border = thin_border

    # Сохраняем файл
    wb.save(filename)
    print(f"\n✅ Готово! Файл создан: {filename}")

if __name__ == "__main__":
    main()