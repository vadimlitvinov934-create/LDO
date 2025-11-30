from datetime import date, datetime, timedelta
from io import BytesIO

from flask import (
    Blueprint,
    render_template,
    request,
    redirect,
    url_for,
    session,
    flash,
    send_file,
)
from core.auth_bp import require_role
from core.permissions import (
    get_head_allowed_prefixes,
    head_list_groups_for_prefixes,
    head_group_allowed,
)
from models import SessionLocal, Student, Attendance, PeriodSkip
from sqlalchemy import func, and_

head_bp = Blueprint("head_bp", __name__, url_prefix="/head")

# ───────────────── helpers ─────────────────


def _today():
    return date.today()


def _parse_day(s: str | None) -> date:
    if not s:
        return _today()
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except Exception:
        return _today()


def _month_range(d: date):
    start = d.replace(day=1)
    if start.month == 12:
        end = start.replace(year=start.year + 1, month=1, day=1) - timedelta(days=1)
    else:
        end = start.replace(month=start.month + 1, day=1) - timedelta(days=1)
    return start, end


# Примитив: семестр — 01.09..31.12 и 01.01..31.05 (можешь поменять под свой колледж)
def _semester_range(d: date):
    if d.month >= 9:  # осенний
        start = date(d.year, 9, 1)
        end = date(d.year, 12, 31)
    else:  # весенний
        start = date(d.year, 1, 1)
        end = date(d.year, 5, 31)
    return start, end


def _attendance_stats(group_code: str, start: date, end: date):
    """Агрегация по статусам в интервале [start..end].

    ВАЖНО: считаем только те пары, которые реально были в расписании
    (если на день для группы есть PeriodSkip по паре, эта пара из статистики
    исключается, даже если по ошибке по ней проставили посещаемость).
    """
    with SessionLocal() as s:
        q = (
            s.query(
                Attendance.status,
                func.count(Attendance.id),
            )
            .join(Student, Student.id == Attendance.student_id)
            .outerjoin(
                PeriodSkip,
                and_(
                    func.date(PeriodSkip.date) == func.date(Attendance.date),
                    PeriodSkip.period_code == Attendance.period_code,
                    func.trim(PeriodSkip.group_code) == func.trim(Student.group_code),
                ),
            )
            .filter(
                func.date(Attendance.date) >= start,
                func.date(Attendance.date) <= end,
                func.trim(Student.group_code) == group_code,
                Attendance.status.in_(["present", "late", "absent", "excused"]),
                PeriodSkip.id.is_(None),  # игнорируем отменённые пары
            )
            .group_by(Attendance.status)
        )

        counts = {"present": 0, "late": 0, "absent": 0, "excused": 0}
        total = 0
        for st, cnt in q.all():
            if st in counts:
                counts[st] += cnt
                total += cnt

        pct = {}
        for k, v in counts.items():
            pct[k] = (v * 100.0 / total) if total else 0.0

        return {"counts": counts, "total": total, "pct": pct}


def _detailed_student_table(group_code: str, start: date, end: date):
    """Таблица по студентам в стиле Excel-прототипа (за период start..end).

    Считаем:
    - всего пропусков в академических часах (по умолчанию 2 часа за пару),
    - разбивку по причинам (заявление / по болезни / соревнования / другое / неуважительная),
    - процент посещаемости по каждому студенту и средний по группе.
    """
    HOURS_PER_LESSON = 2  # если в колледже другая длительность пары — поменяй это значение

    with SessionLocal() as s:
        students = (
            s.query(Student)
            .filter(func.trim(Student.group_code) == group_code)
            .order_by(Student.full_name)
            .all()
        )

        if not students:
            return {"rows": [], "group_pct": 0.0}

        # Берём отметки посещаемости за период, НО сразу исключаем пары, которые были отменены
        records = (
            s.query(Attendance)
            .join(Student, Student.id == Attendance.student_id)
            .outerjoin(
                PeriodSkip,
                and_(
                    func.date(PeriodSkip.date) == func.date(Attendance.date),
                    PeriodSkip.period_code == Attendance.period_code,
                    func.trim(PeriodSkip.group_code) == func.trim(Student.group_code),
                ),
            )
            .filter(
                func.trim(Student.group_code) == group_code,
                func.date(Attendance.date) >= start,
                func.date(Attendance.date) <= end,
                Attendance.status.in_(["present", "absent", "late", "excused"]),
                PeriodSkip.id.is_(None),  # если по паре стоит "таких пар нету" — не учитываем её
            )
            .all()
        )

    # Группируем по студенту
    by_student: dict[int, list[Attendance]] = {}
    for r in records:
        by_student.setdefault(r.student_id, []).append(r)

    rows: list[dict] = []

    for idx, st in enumerate(students, start=1):
        recs = by_student.get(st.id, [])

        total_lessons = len(recs)
        missed_lessons = 0

        by_statement = 0
        by_sick = 0
        by_competition = 0
        by_other = 0
        by_unexcused = 0

        for r in recs:
            status = (r.status or "").strip().lower()
            reason = (r.reason or "").strip().lower()

            # считаем только пропуски (отсутствовал или уважительная)
            if status not in ("absent", "excused"):
                continue

            missed_lessons += 1

            # Грубая эвристика по причинам — под реальные значения можно подстроить
            if "заяв" in reason or reason in ("application",):
                by_statement += 1
            elif "бол" in reason or reason in ("sick",):
                by_sick += 1
            elif "соревн" in reason or reason in ("competition", "contest"):
                by_competition += 1
            elif status == "absent" and not reason:
                # отсутствие без причины → неуважительная
                by_unexcused += 1
            elif status == "absent" and reason:
                # отсутствие с непонятной причиной → тоже считаем как неуважительную,
                # если не попали в категории выше
                by_unexcused += 1
            else:
                # все остальные уважительные
                by_other += 1

        # переведём количество пар в академические часы
        total_hours = missed_lessons * HOURS_PER_LESSON

        row = {
            "num": idx,
            "full_name": st.full_name,
            "total": total_hours,
            "statement": by_statement * HOURS_PER_LESSON,
            "sick": by_sick * HOURS_PER_LESSON,
            "competition": by_competition * HOURS_PER_LESSON,
            "other": by_other * HOURS_PER_LESSON,
            "unexcused": by_unexcused * HOURS_PER_LESSON,
        }

        if total_lessons:
            attended = total_lessons - missed_lessons
            row["pct"] = attended * 100.0 / total_lessons
        else:
            # если по студенту нет отметок — считаем 100% посещаемость
            row["pct"] = 100.0

        rows.append(row)

    group_pct = sum(r["pct"] for r in rows) / len(rows) if rows else 0.0

    return {
        "rows": rows,
        "group_pct": group_pct,
    }


def _load_students(group_code: str):
    with SessionLocal() as s:
        return (
            s.query(Student)
            .filter(func.trim(Student.group_code) == group_code)
            .order_by(Student.full_name)
            .all()
        )


def _load_day_attendance(group_code: str, d: date):
    """Записи посещаемости по группе за день d (для мини-журнала)."""
    with SessionLocal() as s:
        rows = (
            s.query(Attendance)
            .join(Student, Student.id == Attendance.student_id)
            .filter(
                func.trim(Student.group_code) == group_code,
                func.date(Attendance.date) == d,
            )
            .all()
        )
        # Группируем: (student_id -> list of records)
        by_st = {}
        for r in rows:
            by_st.setdefault(r.student_id, []).append(r)
        return by_st


def _load_skips_for_day(group_code: str, d: date):
    with SessionLocal() as s:
        return (
            s.query(PeriodSkip)
            .filter(
                func.date(PeriodSkip.date) == d,
                func.trim(PeriodSkip.group_code) == group_code,
            )
            .all()
        )


# ───────────────── routes ─────────────────


@head_bp.route("/")
@require_role("head")
def choose_group():
    user = session.get("user") or {}
    fio = user.get("fio", "")
    prefixes = get_head_allowed_prefixes(fio)

    if not prefixes:
        flash(
            "Для вашего профиля не настроены префиксы групп. Обратитесь к администратору.",
            "error",
        )
        return render_template(
            "head_groups.html",
            groups=[],
            title="Заведующая — выбор группы",
        )

    groups = head_list_groups_for_prefixes(prefixes)
    return render_template(
        "head_groups.html",
        groups=groups,
        title="Заведующая — выбор группы",
    )


@head_bp.route("/group")
@require_role("head")
def group_view():
    user = session.get("user") or {}
    fio = user.get("fio", "")
    g = (request.args.get("g") or "").strip()
    if not g:
        return redirect(url_for("head_bp.choose_group"))

    if not head_group_allowed(fio, g):
        flash("Эта группа не входит в вашу зону ответственности.", "error")
        return redirect(url_for("head_bp.choose_group"))

    # фильтры
    day = _parse_day(request.args.get("day"))
    mode = (request.args.get("mode") or "day")  # day|month|semester

    if mode == "month":
        start, end = _month_range(day)
    elif mode == "semester":
        start, end = _semester_range(day)
    else:
        start = end = day

    students = _load_students(g)
    day_map = _load_day_attendance(g, day) if mode == "day" else {}
    skips = _load_skips_for_day(g, day)
    stats = _attendance_stats(g, start, end)
    detail = _detailed_student_table(g, start, end)

    return render_template(
        "head_group.html",
        group=g,
        day=day,
        mode=mode,
        students=students,
        day_map=day_map,  # {student_id: [Attendance,...]}
        skips=skips,  # PeriodSkip за день
        stats=stats,  # агрегаты на период
        detail=detail,  # детальная табличка по студентам
        start=start,
        end=end,
        today=_today(),
    )


# ───────────────── выгрузка Excel «Таблица посещаемости по студентам» ─────────────────


@head_bp.route("/group/export_excel")
@require_role("head")
def group_export_excel():
    """
    Выгрузка Excel для таблички «Таблица посещаемости по студентам»
    в формате, максимально похожем на твой файл
    «Посещаемость ПО-175 сентябрь.xlsx».

    Параметры:
        g    — группа
        day  — базовая дата (как в group_view)
        mode — day / month / semester  (за какой период делать отчёт)
    """
    user = session.get("user") or {}
    fio = user.get("fio", "")
    g = (request.args.get("g") or "").strip()
    if not g:
        return redirect(url_for("head_bp.choose_group"))

    if not head_group_allowed(fio, g):
        flash("Эта группа не входит в вашу зону ответственности.", "error")
        return redirect(url_for("head_bp.choose_group"))

    day = _parse_day(request.args.get("day"))
    mode = (request.args.get("mode") or "day")

    if mode == "month":
        start, end = _month_range(day)
        period_title = f"за месяц ({start:%d.%m.%Y}–{end:%d.%m.%Y})"
    elif mode == "semester":
        start, end = _semester_range(day)
        period_title = f"за семестр ({start:%d.%m.%Y}–{end:%d.%m.%Y})"
    else:
        start = end = day
        period_title = f"за {day:%d.%m.%Y}"

    detail = _detailed_student_table(g, start, end)
    rows = detail["rows"]
    group_pct = detail["group_pct"]

    # ─ Excel ─
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Посещаемость"

    bold = Font(bold=True)
    thin = Side(style="thin")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Заголовок
    title = f"Посещаемость группы {g} {period_title}"
    ws.merge_cells("A2:I2")
    cell = ws["A2"]
    cell.value = title
    cell.font = Font(bold=True, size=14)
    cell.alignment = center

    # Шапка таблицы (как в твоём Excel)
    ws["A5"].value = "№"
    ws["B5"].value = "Аты жөні"
    ws["C5"].value = "Всего пропусков(кол-во часов)"
    ws["D5"].value = "В том числе"
    ws["I5"].value = "Процент посещаемости"

    ws["D6"].value = "Заявление"
    ws["E6"].value = "По болезни"
    ws["F6"].value = "Соревнования"
    ws["G6"].value = "Другое"
    ws["H6"].value = "По неуважительной причине"

    # слияния
    ws.merge_cells("A5:A6")
    ws.merge_cells("B5:B6")
    ws.merge_cells("C5:C6")
    ws.merge_cells("D5:H5")
    ws.merge_cells("I5:I6")

    # стили шапки
    for row in (5, 6):
        for col in range(1, 10):
            cell = ws.cell(row=row, column=col)
            cell.font = bold
            cell.alignment = center
            cell.border = border_all

    # ширина колонок
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 22
    ws.column_dimensions["I"].width = 18

    # Данные по студентам
    row_idx = 7
    for r in rows:
        ws.cell(row=row_idx, column=1, value=r["num"]).alignment = center
        ws.cell(row=row_idx, column=2, value=r["full_name"]).alignment = left

        ws.cell(row=row_idx, column=3, value=r["total"]).alignment = center
        ws.cell(row=row_idx, column=4, value=r["statement"]).alignment = center
        ws.cell(row=row_idx, column=5, value=r["sick"]).alignment = center
        ws.cell(row=row_idx, column=6, value=r["competition"]).alignment = center
        ws.cell(row=row_idx, column=7, value=r["other"]).alignment = center
        ws.cell(row=row_idx, column=8, value=r["unexcused"]).alignment = center
        ws.cell(row=row_idx, column=9, value=round(r["pct"], 2)).alignment = center

        for col in range(1, 10):
            ws.cell(row=row_idx, column=col).border = border_all

        row_idx += 1

    # Итоговая строка "Посещаемость по группе"
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=8)
    ws.cell(row=row_idx, column=1, value="Посещаемость по группе").alignment = left
    ws.cell(row=row_idx, column=1).font = bold
    ws.cell(row=row_idx, column=9, value=round(group_pct, 2)).alignment = center
    ws.cell(row=row_idx, column=9).font = bold

    for col in range(1, 10):
        ws.cell(row=row_idx, column=col).border = border_all

    # файл в память
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Посещаемость {g} {start:%Y-%m-%d}_{end:%Y-%m-%d}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
